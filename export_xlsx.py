"""Excel : export (Chantiers + Taches/timeline + Attentes), modele d'import,
et lecture d'un fichier d'import.

Le format d'export et le modele d'import partagent le meme schema (3 feuilles
Chantiers / Taches / Livrables) -> on peut exporter, editer dans Excel,
reimporter.
"""

from __future__ import annotations

import io
import unicodedata
from datetime import date, datetime, timedelta

STATUT_LBL = {"todo": "A faire", "doing": "En cours", "block": "Bloque", "recette": "Recette", "done": "Termine"}
PRIO_LBL = {"h": "Haute", "m": "Moyenne", "b": "Basse"}
LIV_LBL = {"attente": "En attente", "recu": "Recu", "partiel": "Recu partiel", "annule": "Annule"}
RISQUE_LBL = {"ouvert": "Ouvert", "maitrise": "Maitrise", "avere": "Avere", "clos": "Clos"}

PRIO_IN = {"haute": "h", "moyenne": "m", "basse": "b", "h": "h", "m": "m", "b": "b"}
STATUT_IN = {"a faire": "todo", "en cours": "doing", "bloque": "block", "termine": "done",
             "todo": "todo", "doing": "doing", "block": "block", "done": "done"}
LIV_IN = {"en attente": "attente", "attente": "attente", "recu": "recu", "recu partiel": "partiel",
          "partiel": "partiel", "annule": "annule"}
TRUTHY = {"o", "oui", "x", "true", "vrai", "1", "yes", "y"}


def _norm(s) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return s.strip().lower()


def _pct(ch: dict) -> int:
    t = ch.get("taches", [])
    return round(100 * sum(1 for x in t if x["done"]) / len(t)) if t else 0


def _is_late(d) -> bool:
    return bool(d) and d < date.today().isoformat()


def _blocked(ch: dict) -> bool:
    # miroir du frontend isBlocked : bloque uniquement si un livrable non reçu a son échéance
    # dépassée — un livrable rattaché à une tâche ne bloque plus tant que la date n'est pas passée.
    if ch.get("statut") == "done":
        return False
    if (ch.get("blocage") or "").strip():
        return True
    today = date.today().isoformat()
    return any(l.get("statut") in ("attente", "partiel") and l.get("date") and l["date"] < today
               for l in ch.get("livrables", []))   # livrable non reçu et en retard


def _safe(v):
    # anti-injection de formule Excel : un texte commençant par = + - @ (ou tab/CR) devient inerte
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + v
    return v


def _append(ws, row):   # append d'une ligne de données en neutralisant chaque cellule texte
    ws.append([_safe(x) for x in row])


# --------------------------------------------------------------------------- #
# Mini-CPM cote Python (dates planifiees pour la feuille Taches de l'export)
# --------------------------------------------------------------------------- #
def _addd(s: str, n: int) -> str:
    return (datetime.strptime(s, "%Y-%m-%d") + timedelta(days=n)).strftime("%Y-%m-%d")


def _schedule(ch: dict) -> dict:
    tasks = ch.get("taches", [])
    start = ch.get("date_debut") or date.today().isoformat()
    ids = {t["id"] for t in tasks}
    preds = {t["id"]: [p for p in t.get("preds", []) if p in ids and p != t["id"]] for t in tasks}
    dur = lambda t: 0 if t.get("is_milestone") else max(0, int(t.get("duree") or 0))
    ES = {t["id"]: 0 for t in tasks}
    EF = {t["id"]: dur(t) for t in tasks}
    for _ in range(len(tasks) + 1):
        changed = False
        for t in tasks:
            es = max([EF[p] for p in preds[t["id"]]] or [0])
            if es != ES[t["id"]]:
                ES[t["id"]] = es
                EF[t["id"]] = es + dur(t)
                changed = True
        if not changed:
            break
    return {t["id"]: {"start": _addd(start, ES[t["id"]]), "end": _addd(start, EF[t["id"]])} for t in tasks}


# --------------------------------------------------------------------------- #
# Synthèse hebdomadaire (journal + métriques dérivées des dates existantes)
# --------------------------------------------------------------------------- #
def _iso_week_str(ds: str) -> str:
    try:
        y, w, _ = date.fromisoformat(str(ds)[:10]).isocalendar()
        return f"{y}-W{w:02d}"
    except (ValueError, TypeError):
        return ""


def _week_monday(wk: str) -> str:
    try:
        y, w = wk.split("-W")
        return date.fromisocalendar(int(y), int(w), 1).isoformat()
    except (ValueError, TypeError):
        return ""


def _weekly_synthese(store: dict) -> dict:
    from collections import defaultdict
    m = defaultdict(lambda: {"taches": 0, "jalons": 0, "notes": 0, "retours": 0, "relances": 0, "actions": 0})
    for ch in store.get("chantiers", []):
        for t in ch.get("taches", []):
            if t.get("done") and t.get("done_date"):
                w = _iso_week_str(t["done_date"])
                if w:
                    m[w]["taches"] += 1
                    if t.get("is_milestone"):
                        m[w]["jalons"] += 1
        for n in store.get("notes", []):        # l'historique d'un chantier = ses notes
            if n.get("chantier_id") != ch.get("id"):
                continue
            w = _iso_week_str(n.get("date", ""))
            if w:
                m[w]["notes"] += 1
        for l in ch.get("livrables", []):
            w = _iso_week_str(l.get("derniere") or "")
            if w:
                m[w]["relances"] += 1
        for it in ch.get("iterations", []):
            for r in it.get("retours", []):
                w = _iso_week_str(r.get("date") or "")
                if w:
                    m[w]["retours"] += 1
    for j in store.get("journal", []):
        w = j.get("week") or _iso_week_str(j.get("date", ""))
        if w:
            m[w]["actions"] += 1
    return m


# --------------------------------------------------------------------------- #
# Temps chronometre + EVM cote Python (feuilles Temps / EVM de l'export)
# --------------------------------------------------------------------------- #
def _hm2min(h) -> int:   # "HH:MM" -> minutes depuis minuit
    p = str(h or "0:0").split(":")
    try:
        return int(p[0]) * 60 + (int(p[1]) if len(p) > 1 and p[1] != "" else 0)
    except (ValueError, TypeError):
        return 0


def _lunch(store: dict, debut, fin, ds) -> int:
    # minutes de pause dejeuner a exclure (aucune le vendredi) : miroir de lunchOverlap() du front
    try:
        if date.fromisoformat(str(ds)[:10]).weekday() == 4:
            return 0
    except (ValueError, TypeError):
        pass
    stg = store.get("settings", {})
    pd, pf = stg.get("pause_debut"), stg.get("pause_fin")
    if not pd or not pf or pf <= pd:
        return 0
    s, e = _hm2min(debut), _hm2min(fin)
    if e < s:
        e += 1440
    return max(0, min(e, _hm2min(pf)) - max(s, _hm2min(pd)))


def _sess_min(store: dict, s: dict) -> int:
    # duree d'une session chrono en minutes : miroir de sessMin() du front (borne fin de journee,
    # deduit la pause dejeuner). store.load() a deja ferme les chronos oublies des jours passes.
    from store import _day_end, _hm
    today = date.today().isoformat()
    e = s.get("fin")
    if not e:   # session encore active : borne a la fin de journee de sa date
        de = _day_end(store, s.get("date") or today)
        e = de if (s.get("date") and s["date"] < today) else min(_hm(), de)
    m = _hm2min(e) - _hm2min(s.get("debut"))
    if m < 0:
        m += 1440
    return max(0, m - _lunch(store, s.get("debut"), e, s.get("date")))


def _chantier_min(store: dict, cid: str) -> int:
    return sum(_sess_min(store, s) for s in store.get("timelog", []) if s.get("chantier_id") == cid)


def _taux_heure(store: dict) -> float:   # taux journalier / heures facturables ; 15 EUR/h par defaut (comme le front)
    stg = store.get("settings", {})
    t = float(stg.get("taux_jour") or 0)
    hj = float(stg.get("heures_jour") or 7) or 7
    return t / hj if t else 15.0


def _dbetween(a: str, b: str) -> int:
    try:
        return (date.fromisoformat(str(b)[:10]) - date.fromisoformat(str(a)[:10])).days
    except (ValueError, TypeError):
        return 0


def _evm(store: dict, ch: dict, sched: dict) -> dict:
    # Port Python de la fonction evm() du front, sur le planning _schedule (calendaire) deja calcule.
    # PV/EV = BAC pondere par la duree planifiee ; AC = jours-personnes chronometres x taux journalier.
    bac = ch.get("budget")
    BAC = float(bac) if bac not in (None, "") else None
    tasks = [t for t in ch.get("taches", []) if not t.get("is_milestone")]
    Wtot = sum(max(1, t.get("duree") or 1) for t in tasks) or 1
    today = date.today().isoformat()
    pv_frac = ev_frac = 0.0
    for t in tasks:
        w = max(1, t.get("duree") or 1) / Wtot
        sc = sched.get(t["id"])
        if sc:
            if today >= sc["end"]:
                frac = 1.0
            elif today > sc["start"]:
                tot = max(1, _dbetween(sc["start"], sc["end"]))
                frac = min(1.0, _dbetween(sc["start"], today) / tot)
            else:
                frac = 0.0
            pv_frac += w * frac
        if t.get("done"):
            prog = 1.0
        elif t.get("start_date"):
            subs = t.get("subtasks") or []
            prog = (sum(1 for x in subs if x.get("done")) / len(subs)) if subs else 0.5
        else:
            prog = 0.0
        ev_frac += w * prog
    PV = BAC * pv_frac if BAC is not None else None
    EV = BAC * ev_frac if BAC is not None else None
    taux = float((store.get("settings") or {}).get("taux_jour") or 0)
    hj = float((store.get("settings") or {}).get("heures_jour") or 7) or 7
    pjours = _chantier_min(store, ch["id"]) / (hj * 60)   # jours-personnes chronometres
    AC = pjours * taux if taux else None
    SPI = (EV / PV) if (PV and EV is not None) else None
    CPI = (EV / AC) if (AC and EV is not None) else None
    SV = (EV - PV) if (PV is not None and EV is not None) else None
    CV = (EV - AC) if (AC is not None and EV is not None) else None
    EAC = (BAC / CPI) if (BAC is not None and CPI) else None
    VAC = (BAC - EAC) if (BAC is not None and EAC is not None) else None
    return {"BAC": BAC, "PV": PV, "EV": EV, "AC": AC, "SPI": SPI, "CPI": CPI,
            "SV": SV, "CV": CV, "EAC": EAC, "VAC": VAC, "pjours": pjours}


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #
def _styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        "head_fill": PatternFill("solid", fgColor="2D2B28"),
        "head_font": Font(color="FFFFFF", bold=True),
        "late_font": Font(color="A8261B", bold=True),
        "border": Border(*[Side(style="thin", color="D9D7CF")] * 4),
        "Alignment": Alignment,
    }


def _header(ws, ncols, st):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill, c.font = st["head_fill"], st["head_font"]
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20


def _widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w


def build() -> bytes:
    from openpyxl import Workbook
    from store import load

    store = load()
    st = _styles()
    Align = st["Alignment"]
    wb = Workbook()

    # --- Chantiers ---
    ws = wb.active
    ws.title = "Chantiers"
    cols = ["Titre", "Objectif", "Debut", "Echeance", "Priorite", "Statut", "Theme",
            "Avancement %", "Point bloquant"]
    ws.append(cols)
    th_nom = {t["id"]: t.get("nom", "") for t in store.get("themes", [])}
    for ch in sorted(store["chantiers"], key=lambda c: c.get("ordre", 0)):
        _append(ws, [ch["titre"], ch.get("objectif") or "", ch.get("date_debut") or "",
                   ch.get("echeance") or "", PRIO_LBL.get(ch["prio"], ch["prio"]),
                   "Bloque" if _blocked(ch) else STATUT_LBL.get(ch["statut"], ch["statut"]),
                   th_nom.get(ch.get("theme_id"), ""), _pct(ch), ch.get("blocage") or ""])
        r = ws.max_row
        if ch["statut"] != "done" and _is_late(ch.get("echeance")):
            ws.cell(row=r, column=4).font = st["late_font"]
        for col in range(1, len(cols) + 1):
            ws.cell(row=r, column=col).border = st["border"]
            ws.cell(row=r, column=col).alignment = Align(vertical="top", wrap_text=col in (2, 9))
    _widths(ws, [30, 44, 12, 12, 10, 10, 22, 12, 36])
    _header(ws, len(cols), st)

    # --- Taches (plan + timeline) ---
    ws2 = wb.create_sheet("Taches")
    cols2 = ["Chantier", "Tache", "Jalon", "Duree (j)", "Predecesseurs", "Fait",
             "Debut prevu", "Fin prevue"]
    ws2.append(cols2)
    for ch in sorted(store["chantiers"], key=lambda c: c.get("ordre", 0)):
        sched = _schedule(ch)
        lbl = {t["id"]: t["label"] for t in ch["taches"]}
        for t in ch["taches"]:
            preds = "; ".join(lbl.get(p, "") for p in t.get("preds", []) if p in lbl)
            _append(ws2, [ch["titre"], t["label"], "O" if t.get("is_milestone") else "",
                        0 if t.get("is_milestone") else t.get("duree", 1), preds,
                        "O" if t["done"] else "", sched[t["id"]]["start"], sched[t["id"]]["end"]])
            for col in range(1, len(cols2) + 1):
                ws2.cell(row=ws2.max_row, column=col).border = st["border"]
    _widths(ws2, [28, 38, 7, 10, 30, 6, 12, 12])
    _header(ws2, len(cols2), st)

    # --- Attentes (par personne) ---
    ws3 = wb.create_sheet("Attentes")
    cols3 = ["Personne", "Role", "Chantier", "Livrable attendu", "Pour le", "Statut", "Relances", "Impact"]
    ws3.append(cols3)
    rows = []
    for ch in store["chantiers"]:
        for l in ch["livrables"]:
            late = l["statut"] == "attente" and _is_late(l.get("date"))
            rows.append((l["personne"], l.get("role", ""), ch["titre"], l["quoi"],
                         l.get("date") or "", "EN RETARD" if late else LIV_LBL.get(l["statut"], l["statut"]),
                         l.get("relances", 0), l.get("impact") or "", late))
    rows.sort(key=lambda r: (r[0].lower(), r[4] or "9999"))
    for r in rows:
        _append(ws3, list(r[:8]))
        if r[8]:
            ws3.cell(row=ws3.max_row, column=6).font = st["late_font"]
        for col in range(1, len(cols3) + 1):
            ws3.cell(row=ws3.max_row, column=col).border = st["border"]
            ws3.cell(row=ws3.max_row, column=col).alignment = Align(vertical="top", wrap_text=col in (4, 8))
    _widths(ws3, [22, 16, 28, 40, 12, 13, 9, 34])
    _header(ws3, len(cols3), st)

    # --- Synthese hebdo (progression semaine par semaine) ---
    ws4 = wb.create_sheet("Synthese hebdo")
    cols4 = ["Semaine", "Debut (lundi)", "Taches terminees", "Jalons", "Notes",
             "Retours recus", "Relances", "Actions"]
    ws4.append(cols4)
    syn = _weekly_synthese(store)
    for w in sorted(syn.keys(), reverse=True):
        d = syn[w]
        ws4.append([w, _week_monday(w), d["taches"], d["jalons"], d["notes"],
                    d["retours"], d["relances"], d["actions"]])
        for col in range(1, len(cols4) + 1):
            ws4.cell(row=ws4.max_row, column=col).border = st["border"]
    _widths(ws4, [12, 14, 16, 9, 9, 14, 10, 9])
    _header(ws4, len(cols4), st)

    # --- Journal (toutes les actions horodatees) ---
    ws5 = wb.create_sheet("Journal")
    cols5 = ["Date", "Semaine", "Chantier", "Action"]
    ws5.append(cols5)
    for j in sorted(store.get("journal", []), key=lambda x: x.get("ts", ""), reverse=True):
        _append(ws5, [j.get("date", ""), j.get("week", ""), j.get("chantier", ""), j.get("msg", "")])
        for col in range(1, len(cols5) + 1):
            ws5.cell(row=ws5.max_row, column=col).border = st["border"]
            ws5.cell(row=ws5.max_row, column=col).alignment = Align(vertical="top", wrap_text=col == 4)
    _widths(ws5, [12, 12, 28, 64])
    _header(ws5, len(cols5), st)

    # --- Cahiers des charges (synthese + validation) ---
    CDC_LBL = {"brouillon": "Brouillon", "en_validation": "En validation",
               "valide": "Valide", "obsolete": "Obsolete"}
    ws6 = wb.create_sheet("Cahiers des charges")
    cols6 = ["Chantier", "Reference", "Titre", "Indice", "Statut", "Redige par",
             "Mis a jour", "Valide par", "Date validation", "Parties prenantes", "Lien"]
    ws6.append(cols6)
    for ch in store.get("chantiers", []):
        cdc = ch.get("cdc")
        if not cdc:
            continue
        pp = " ; ".join(f"{p.get('nom', '')}{(' (' + p['role'] + ')') if p.get('role') else ''}"
                        for p in cdc.get("parties_prenantes", []))
        _append(ws6, [ch.get("titre", ""), cdc.get("reference", ""), cdc.get("titre", ""),
                    cdc.get("indice", ""), CDC_LBL.get(cdc.get("statut"), cdc.get("statut", "")),
                    cdc.get("redacteur", ""), cdc.get("date_maj", ""), cdc.get("valide_par", ""),
                    cdc.get("date_validation", "") or "", pp, cdc.get("lien", "")])
        for col in range(1, len(cols6) + 1):
            ws6.cell(row=ws6.max_row, column=col).border = st["border"]
            ws6.cell(row=ws6.max_row, column=col).alignment = Align(vertical="top", wrap_text=col in (3, 10))
    _widths(ws6, [24, 16, 30, 7, 13, 12, 12, 16, 13, 34, 28])
    _header(ws6, len(cols6), st)

    # --- CdC : suivi des modifications (revisions, tous chantiers) ---
    ws7 = wb.create_sheet("CdC revisions")
    cols7 = ["Chantier", "Indice", "Date", "Auteur", "Objet de la modification"]
    ws7.append(cols7)
    for ch in store.get("chantiers", []):
        cdc = ch.get("cdc")
        if not cdc:
            continue
        for r in cdc.get("revisions", []):
            ws7.append([ch.get("titre", ""), r.get("indice", ""), r.get("date", ""),
                        r.get("auteur", ""), r.get("objet", "")])
            for col in range(1, len(cols7) + 1):
                ws7.cell(row=ws7.max_row, column=col).border = st["border"]
                ws7.cell(row=ws7.max_row, column=col).alignment = Align(vertical="top", wrap_text=col == 5)
    _widths(ws7, [24, 7, 12, 16, 60])
    _header(ws7, len(cols7), st)

    # --- Risques (registre proba x gravite 5x5, tous chantiers) ---
    ws8 = wb.create_sheet("Risques")
    cols8 = ["Chantier", "Risque", "Categorie", "Probabilite", "Gravite", "Criticite",
             "Parade", "Responsable", "Revue le", "Statut"]
    ws8.append(cols8)
    for ch in store.get("chantiers", []):
        for rk in ch.get("risques", []):
            crit = (rk.get("probabilite") or 0) * (rk.get("gravite") or 0)
            _append(ws8, [ch["titre"], rk.get("libelle", ""), rk.get("categorie", ""),
                        rk.get("probabilite", ""), rk.get("gravite", ""), crit,
                        rk.get("parade", ""), rk.get("responsable", ""), rk.get("echeance_revue") or "",
                        RISQUE_LBL.get(rk.get("statut"), rk.get("statut", ""))])
            r = ws8.max_row
            if crit >= 15 and rk.get("statut") != "clos":   # criticite elevee non close : mise en evidence
                ws8.cell(row=r, column=6).font = st["late_font"]
            for col in range(1, len(cols8) + 1):
                ws8.cell(row=r, column=col).border = st["border"]
                ws8.cell(row=r, column=col).alignment = Align(vertical="top", wrap_text=col in (2, 7))
    _widths(ws8, [26, 34, 16, 11, 8, 9, 40, 18, 12, 10])
    _header(ws8, len(cols8), st)

    # --- EVM (valeur acquise, une ligne par chantier budgete) ---
    ws9 = wb.create_sheet("EVM")
    cols9 = ["Chantier", "Budget (BAC)", "Valeur planifiee (PV)", "Valeur acquise (EV)",
             "Cout reel (AC)", "SPI", "CPI", "Ecart delai (SV)", "Ecart cout (CV)",
             "Cout final (EAC)", "Ecart final (VAC)", "Jours-pers"]
    ws9.append(cols9)
    _eur = lambda v: round(v) if v is not None else ""
    _idx = lambda v: round(v, 2) if v is not None else ""
    for ch in sorted(store["chantiers"], key=lambda c: c.get("ordre", 0)):
        E = _evm(store, ch, _schedule(ch))
        if E["BAC"] is None:   # seuls les chantiers avec un budget defini
            continue
        _append(ws9, [ch["titre"], _eur(E["BAC"]), _eur(E["PV"]), _eur(E["EV"]), _eur(E["AC"]),
                    _idx(E["SPI"]), _idx(E["CPI"]), _eur(E["SV"]), _eur(E["CV"]),
                    _eur(E["EAC"]), _eur(E["VAC"]), round(E["pjours"], 1)])
        r = ws9.max_row
        if E["SV"] is not None and E["SV"] < 0:      # retard planning
            ws9.cell(row=r, column=8).font = st["late_font"]
        if E["VAC"] is not None and E["VAC"] < 0:    # depassement de budget prevu
            ws9.cell(row=r, column=11).font = st["late_font"]
        for col in range(1, len(cols9) + 1):
            ws9.cell(row=r, column=col).border = st["border"]
    _widths(ws9, [28, 13, 16, 15, 13, 7, 7, 13, 13, 14, 14, 10])
    _header(ws9, len(cols9), st)

    # --- Temps (timelog agrege par chantier et par type, valorise au taux horaire) ---
    from collections import defaultdict
    KIND_LBL = {"tache": "Tache", "action": "Action / routine", "recette": "Recette", "libre": "Libre"}
    ws10 = wb.create_sheet("Temps")
    cols10 = ["Chantier", "Type", "Sessions", "Temps (h)", "Cout (EUR)"]
    ws10.append(cols10)
    ch_titre = {c["id"]: c["titre"] for c in store["chantiers"]}
    agg = defaultdict(lambda: {"n": 0, "min": 0})
    for s in store.get("timelog", []):
        key = (ch_titre.get(s.get("chantier_id"), "(sans chantier)"), s.get("kind") or "libre")
        agg[key]["n"] += 1
        agg[key]["min"] += _sess_min(store, s)
    th = _taux_heure(store)
    for titre, kind in sorted(agg.keys(), key=lambda k: (k[0].lower(), k[1])):
        d = agg[(titre, kind)]
        _append(ws10, [titre, KIND_LBL.get(kind, kind), d["n"],
                     round(d["min"] / 60, 2), round(d["min"] / 60 * th)])
        for col in range(1, len(cols10) + 1):
            ws10.cell(row=ws10.max_row, column=col).border = st["border"]
    _widths(ws10, [28, 14, 10, 10, 12])
    _header(ws10, len(cols10), st)

    # -- Actions : taches libres et routines, avec le taux de tenue ----------
    ws11 = wb.create_sheet("Actions")
    cols11 = ["Libelle", "Type", "Theme", "Chantier", "Priorite", "Echeance",
              "Etat", "Recurrence", "Tenue (%)", "Fait / du", "Temps (h)"]
    ws11.append(cols11)
    th_nom = {t["id"]: t.get("nom", "") for t in store.get("themes", [])}
    tl_min = defaultdict(int)
    for s in store.get("timelog", []):
        if s.get("action_id"):
            tl_min[s["action_id"]] += _sess_min(store, s)
    for a in store.get("actions", []):
        rec = a.get("recurrence")
        occ = a.get("occurrences", [])
        faits = sum(1 for o in occ if o.get("statut") == "fait")
        # denominateur : les occurrences enregistrees hors sautees volontairement
        base = sum(1 for o in occ if o.get("statut") in ("fait", "rate"))
        if rec:
            freq = {"jour": "Chaque jour", "semaine": "Chaque semaine", "mois": "Chaque mois"}.get(rec["freq"], rec["freq"])
            if rec["freq"] == "semaine" and rec.get("jours"):
                freq += " (" + ", ".join(["lun", "mar", "mer", "jeu", "ven", "sam", "dim"][j] for j in rec["jours"]) + ")"
            if rec["freq"] == "mois":
                freq += " (dernier jour)" if rec.get("jour_mois") == "fin" else f" (le {rec.get('jour_mois') or 1})"
            etat = "Active" if a.get("actif") else "En sommeil"
        else:
            freq, etat = "", ("Faite" if a.get("done") else "A faire")
        _append(ws11, [a.get("label", ""), "Routine" if rec else "Action",
                       th_nom.get(a.get("theme_id"), ""),
                       ch_titre.get(a.get("chantier_id"), ""),
                       PRIO_LBL.get(a.get("prio"), a.get("prio", "")),
                       a.get("echeance") or "", etat, freq,
                       (round(100 * faits / base) if base else ""),
                       (f"{faits}/{base}" if base else ""),
                       round(tl_min.get(a["id"], 0) / 60, 2)])
        for col in range(1, len(cols11) + 1):
            ws11.cell(row=ws11.max_row, column=col).border = st["border"]
    _widths(ws11, [38, 10, 22, 26, 10, 12, 12, 24, 10, 11, 10])
    _header(ws11, len(cols11), st)

    # -- Notes : le journal horodate, exportable tel quel --------------------
    ws12 = wb.create_sheet("Notes")
    cols12 = ["Date", "Heure", "Type", "Titre", "Theme", "Chantier", "Contenu", "Modifiee le"]
    ws12.append(cols12)
    nt_lbl = {"note": "Note", "reunion": "Reunion", "decision": "Decision", "idee": "Idee"}
    for n in sorted(store.get("notes", []),
                    key=lambda x: (x.get("date") or "", x.get("heure") or ""), reverse=True):
        _append(ws12, [n.get("date") or "", n.get("heure") or "",
                       nt_lbl.get(n.get("type"), n.get("type", "")),
                       n.get("titre") or "", th_nom.get(n.get("theme_id"), ""),
                       ch_titre.get(n.get("chantier_id"), ""),
                       n.get("corps") or "", n.get("maj_le") or ""])
        for col in range(1, len(cols12) + 1):
            c12 = ws12.cell(row=ws12.max_row, column=col)
            c12.border = st["border"]
            if col == 7:
                c12.alignment = Align(wrap_text=True, vertical="top")
    _widths(ws12, [11, 8, 11, 26, 22, 26, 70, 12])
    _header(ws12, len(cols12), st)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Modele d'import
# --------------------------------------------------------------------------- #
def build_template() -> bytes:
    from openpyxl import Workbook
    st = _styles()
    wb = Workbook()

    ws = wb.active
    ws.title = "Chantiers"
    ws.append(["Titre", "Objectif", "Debut", "Echeance", "Priorite", "Statut", "Theme"])
    ws.append(["Nouveau site web", "Refondre le site vitrine", "2026-06-10", "2026-07-15",
               "Haute", "En cours", "SI, Infra & Collaboration"])
    _widths(ws, [28, 40, 12, 12, 10, 10, 20]); _header(ws, 7, st)

    ws2 = wb.create_sheet("Taches")
    ws2.append(["Chantier", "Tache", "Jalon", "Duree (j)", "Predecesseurs"])
    ex = [
        ("Nouveau site web", "Cadrage", "", 2, ""),
        ("Nouveau site web", "Maquettes", "", 4, "Cadrage"),
        ("Nouveau site web", "Validation maquettes", "O", 0, "Maquettes"),
        ("Nouveau site web", "Integration", "", 6, "Validation maquettes"),
        ("Nouveau site web", "Contenus", "", 5, "Cadrage"),
        ("Nouveau site web", "Recette", "", 3, "Integration; Contenus"),
        ("Nouveau site web", "Mise en ligne", "O", 0, "Recette"),
    ]
    for r in ex:
        ws2.append(list(r))
    _widths(ws2, [28, 34, 7, 10, 34]); _header(ws2, 5, st)

    ws3 = wb.create_sheet("Livrables")
    ws3.append(["Chantier", "Personne", "Role", "Livrable", "Pour le", "Statut", "Impact"])
    ws3.append(["Nouveau site web", "Agence", "Prestataire", "Livrer les maquettes",
                "2026-06-20", "En attente", "Bloque l'integration"])
    _widths(ws3, [28, 20, 16, 36, 12, 13, 30]); _header(ws3, 7, st)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Lecture d'un fichier d'import
# --------------------------------------------------------------------------- #
def _sheet(wb, *names):
    want = {_norm(n) for n in names}
    for s in wb.worksheets:
        if _norm(s.title) in want:
            return s
    return None


def _rows(ws):
    """Renvoie une liste de dicts {header_normalise: valeur}."""
    if ws is None:
        return []
    it = ws.iter_rows(values_only=True)
    try:
        headers = [_norm(h) for h in next(it)]
    except StopIteration:
        return []
    out = []
    for row in it:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        out.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    return out


def _cell(d, *keys):
    for k in keys:
        if k in d and d[k] is not None and str(d[k]).strip() != "":
            return d[k]
    return None


def _to_date(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    return s[:10] if len(s) >= 10 and s[4] == "-" else s


def parse_import(data: bytes) -> dict:
    """Lit un classeur d'import et renvoie une structure normalisee :
    {chantiers:[...], taches:{key:[...]}, livrables:{key:[...]}}.
    `key` = titre normalise (pour relier les feuilles entre elles)."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(data), data_only=True, read_only=True)

    chantiers, taches, livrables = [], {}, {}
    seen = {}

    def ensure(titre):
        k = _norm(titre)
        if k not in seen:
            seen[k] = {"key": k, "titre": str(titre).strip(), "objectif": "", "debut": None,
                       "echeance": None, "prio": "m", "statut": "todo", "theme": ""}
            chantiers.append(seen[k])
            taches[k] = []
            livrables[k] = []
        return seen[k]

    for d in _rows(_sheet(wb, "Chantiers", "Chantier")):
        titre = _cell(d, "titre", "chantier")
        if not titre:
            continue
        c = ensure(titre)
        c["objectif"] = str(_cell(d, "objectif") or "")
        c["debut"] = _to_date(_cell(d, "debut", "date debut", "debut prevu"))
        c["echeance"] = _to_date(_cell(d, "echeance", "echeance prevue"))
        c["prio"] = PRIO_IN.get(_norm(_cell(d, "priorite", "priorite")), "m")
        c["statut"] = STATUT_IN.get(_norm(_cell(d, "statut")), "todo")
        # Le theme est choisi dans la liste FERMEE : un libelle inconnu est ignore
        # (l'import ne doit pas pouvoir recreer un fourre-tout de tags libres).
        c["theme"] = str(_cell(d, "theme", "thème", "tags") or "").strip()

    for d in _rows(_sheet(wb, "Taches", "Tache")):
        titre = _cell(d, "chantier")
        label = _cell(d, "tache", "tâche", "label")
        if not titre or not label:
            continue
        ensure(titre)
        preds = _cell(d, "predecesseurs", "predecesseur", "preds")
        preds = [p.strip() for p in str(preds).replace(",", ";").split(";") if p.strip()] if preds else []
        jalon = _norm(_cell(d, "jalon")) in TRUTHY
        try:
            duree = 0 if jalon else max(0, int(float(_cell(d, "duree", "duree (j)", "duree j") or 1)))
        except (ValueError, TypeError):
            duree = 1
        taches[_norm(titre)].append({"label": str(label).strip(), "is_milestone": jalon,
                                     "duree": duree, "preds": preds})

    for d in _rows(_sheet(wb, "Livrables", "Livrable", "Attentes")):
        titre = _cell(d, "chantier")
        quoi = _cell(d, "livrable", "livrable attendu", "quoi")
        personne = _cell(d, "personne")
        if not titre or not quoi or not personne:
            continue
        ensure(titre)
        livrables[_norm(titre)].append({
            "personne": str(personne).strip(), "role": str(_cell(d, "role") or ""),
            "quoi": str(quoi).strip(), "date": _to_date(_cell(d, "pour le", "date", "echeance")),
            "statut": LIV_IN.get(_norm(_cell(d, "statut")), "attente"),
            "impact": str(_cell(d, "impact") or "")})

    wb.close()
    return {"chantiers": chantiers, "taches": taches, "livrables": livrables}


def import_into(store: dict, data: bytes):
    """Cree de NOUVEAUX chantiers dans `store` a partir du classeur. Renvoie
    (nb_chantiers, nb_taches, nb_livrables). Mute store ; l'appelant sauvegarde."""
    import store as store_mod
    parsed = parse_import(data)
    n_ch = n_t = n_l = 0
    for c in parsed["chantiers"]:
        store_mod.apply_op(store, {"op": "create_chantier", "titre": c["titre"],
                                   "objectif": c["objectif"], "prio": c["prio"], "statut": c["statut"],
                                   "echeance": c["echeance"], "date_debut": c["debut"]})
        ch = store["chantiers"][-1]
        n_ch += 1
        if c.get("theme"):   # rapproche du theme existant ; sinon le chantier reste sans theme
            tid = next((t["id"] for t in store.get("themes", [])
                        if _norm(t.get("nom", "")) == _norm(c["theme"])), None)
            if tid:
                store_mod.apply_op(store, {"op": "set_theme", "chantier_id": ch["id"], "theme_id": tid})
        # taches (1re passe : creation, on garde label -> id)
        label_to_id = {}
        for t in parsed["taches"].get(c["key"], []):
            store_mod.apply_op(store, {"op": "add_tache", "chantier_id": ch["id"], "label": t["label"],
                                       "is_milestone": t["is_milestone"], "duree": t["duree"]})
            label_to_id[_norm(t["label"])] = ch["taches"][-1]["id"]
            n_t += 1
        # 2e passe : resolution des predecesseurs par label
        for t in parsed["taches"].get(c["key"], []):
            if t["preds"]:
                pred_ids = [label_to_id[_norm(p)] for p in t["preds"] if _norm(p) in label_to_id]
                if pred_ids:
                    tid = label_to_id[_norm(t["label"])]
                    store_mod.apply_op(store, {"op": "update_tache", "chantier_id": ch["id"],
                                               "tache_id": tid, "preds": pred_ids})
        for l in parsed["livrables"].get(c["key"], []):
            store_mod.apply_op(store, {"op": "add_livrable", "chantier_id": ch["id"],
                                       "personne": l["personne"], "role": l["role"], "quoi": l["quoi"],
                                       "date": l["date"], "statut": l["statut"], "impact": l["impact"]})
            n_l += 1
    return n_ch, n_t, n_l
